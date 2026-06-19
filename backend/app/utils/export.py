"""
Export weather records to CSV, Excel (.xlsx), or JSON.

Each export flattens the stored JSON into one row per day so the
spreadsheet / CSV is immediately useful in any data tool.
"""

import csv
import io
import json
from datetime import datetime
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.models.weather import WeatherRecord

ExportFormat = Literal["csv", "excel", "json"]

_HEADERS = [
    "record_id",
    "label",
    "location_query",
    "location_name",
    "country_code",
    "latitude",
    "longitude",
    "record_date_from",
    "record_date_to",
    "date",
    "temp_max_c",
    "temp_min_c",
    "temp_mean_c",
    "precipitation_mm",
    "windspeed_max_kmh",
    "weather_code",
    "weather_description",
    "precipitation_probability_pct",
    "uv_index_max",
    "notes",
    "created_at",
]


def _flatten_record(record: WeatherRecord) -> list[dict]:
    """Turn one WeatherRecord into one dict per daily entry."""
    base = {
        "record_id": record.id,
        "label": record.label or "",
        "location_query": record.location_query,
        "location_name": record.location_name,
        "country_code": record.country_code or "",
        "latitude": record.latitude,
        "longitude": record.longitude,
        "record_date_from": record.date_from.isoformat(),
        "record_date_to": record.date_to.isoformat(),
        "notes": record.notes or "",
        "created_at": record.created_at.strftime("%Y-%m-%d %H:%M:%S")
        if record.created_at
        else "",
    }

    daily = record.weather_data.get("daily", [])
    if not daily:
        return [{**base, **{h: "" for h in _HEADERS if h not in base}}]

    rows = []
    for day in daily:
        rows.append(
            {
                **base,
                "date": day.get("date", ""),
                "temp_max_c": day.get("temp_max", ""),
                "temp_min_c": day.get("temp_min", ""),
                "temp_mean_c": day.get("temp_mean", ""),
                "precipitation_mm": day.get("precipitation_mm", ""),
                "windspeed_max_kmh": day.get("windspeed_max_kmh", ""),
                "weather_code": day.get("weather_code", ""),
                "weather_description": day.get("weather_description", ""),
                "precipitation_probability_pct": day.get(
                    "precipitation_probability_pct", ""
                ),
                "uv_index_max": day.get("uv_index_max", ""),
            }
        )
    return rows


def _to_csv(records: list[WeatherRecord]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_HEADERS, extrasaction="ignore")
    writer.writeheader()
    for record in records:
        writer.writerows(_flatten_record(record))
    return buf.getvalue().encode("utf-8-sig")  # BOM so Excel opens it correctly


def _to_excel(records: list[WeatherRecord]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Weather Records"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2C5F8A")
    center = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    row_num = 2
    for record in records:
        for flat in _flatten_record(record):
            for col, key in enumerate(_HEADERS, start=1):
                ws.cell(row=row_num, column=col, value=flat.get(key, ""))
            row_num += 1

    # Auto-size columns (capped at 40 chars wide)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _to_json(records: list[WeatherRecord]) -> bytes:
    output = []
    for record in records:
        output.append(
            {
                "id": record.id,
                "label": record.label,
                "location_query": record.location_query,
                "location_name": record.location_name,
                "country_code": record.country_code,
                "latitude": record.latitude,
                "longitude": record.longitude,
                "date_from": record.date_from.isoformat(),
                "date_to": record.date_to.isoformat(),
                "weather": record.weather_data,
                "notes": record.notes,
                "created_at": record.created_at.isoformat()
                if record.created_at
                else None,
            }
        )
    return json.dumps(output, indent=2).encode("utf-8")


def export_records(
    records: list[WeatherRecord], format: ExportFormat
) -> tuple[bytes, str, str]:
    """
    Serialize records to the requested format.

    Returns (content_bytes, media_type, filename).
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if format == "csv":
        return _to_csv(records), "text/csv", f"weather_export_{timestamp}.csv"

    if format == "excel":
        return (
            _to_excel(records),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"weather_export_{timestamp}.xlsx",
        )

    return _to_json(records), "application/json", f"weather_export_{timestamp}.json"
